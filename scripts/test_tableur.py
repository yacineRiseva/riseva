#!/usr/bin/env python3
"""Le classeur, relu par un lecteur qui n'est pas le nôtre.

Un écrivain de `.xlsx` écrit à la main produit très facilement une archive
parfaitement crédible qu'Excel ouvre vide, sans un mot d'explication. Le seul
test qui vaille est un aller-retour : on fabrique le classeur avec le code de
production, et on le relit avec openpyxl, qui n'a rien à voir avec nous.

Ce qu'on vérifie, et pourquoi :
  — les onglets portent le nom des rubriques, accents compris ;
  — un nombre est un nombre, pas une chaîne : un total qui ne s'additionne pas
    est le premier reproche qu'on fait à un export ;
  — une valeur absente reste vide, jamais zéro : les deux ne se distinguent
    plus ensuite, et c'est la seule erreur de ce fichier qui se retourne contre
    celui qui l'a publié ;
  — la première ligne est bien l'en-tête, en gras.

    python3 scripts/test_tableur.py
"""
import subprocess, sys, pathlib, tempfile, shutil

RACINE = pathlib.Path(__file__).resolve().parent.parent
SORTIE = pathlib.Path(tempfile.gettempdir()) / "riseva-classeur.xlsx"

FABRIQUE = """
import { DB } from "./data.js";
import { classeur } from "./tableur.js";
import { writeFileSync } from "node:fs";
const c = DB.campagnes().filter(x => x.etat === "ouverte")[0]
       || DB.campagnes()[0];
const onglets = DB.classeurCollecte(c.id);
const blob = classeur(onglets);
writeFileSync(process.argv[2], Buffer.from(await blob.arrayBuffer()));
process.stdout.write(JSON.stringify({
  campagne: c.libelle,
  onglets: onglets.map(o => o.nom),
  lignes: onglets.map(o => o.lignes.length)
}));
"""

def main():
    tmp = pathlib.Path(tempfile.mkdtemp())
    for f in ("qr.js", "tableur.js", "data.js"):
        shutil.copy(RACINE / "public" / "app" / f, tmp / f)
    (tmp / "fabrique.mjs").write_text(FABRIQUE, encoding="utf-8")
    r = subprocess.run(["node", str(tmp / "fabrique.mjs"), str(SORTIE)],
                       capture_output=True, text=True, cwd=tmp)
    if r.returncode:
        print("  RATÉ le classeur n'a pas pu être fabriqué")
        print("      " + r.stderr.strip()[:500]); return 1
    import json
    meta = json.loads(r.stdout)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  À FAIRE openpyxl absent : le classeur n'a pas été relu.")
        print("          pip install openpyxl --break-system-packages")
        return 0

    rates = []
    def dit(nom, ok):
        print(("  ok   " if ok else "  RATÉ ") + nom)
        if not ok: rates.append(nom)

    wb = load_workbook(SORTIE)
    dit("le classeur s'ouvre avec un lecteur indépendant", True)
    dit("les onglets portent le nom des rubriques, accents compris",
        wb.sheetnames == meta["onglets"])
    dit("les définitions ferment le classeur, elles ne l'ouvrent pas",
        wb.sheetnames[-1] == "Définitions" and wb.sheetnames[0] != "Définitions")

    f = wb[wb.sheetnames[0]]
    dit("la première ligne est l'en-tête, en gras",
        f.cell(1, 1).value == "Site" and bool(f.cell(1, 1).font.bold))
    nombres = [c.value for ligne in f.iter_rows(min_row=2) for c in ligne
               if isinstance(c.value, (int, float))]
    dit("les nombres sont des nombres, pas des chaînes", len(nombres) > 0)
    textes_numeriques = [c.value for ligne in f.iter_rows(min_row=2) for c in ligne
                         if isinstance(c.value, str)
                         and c.value.replace(".", "", 1).isdigit()]
    dit("aucun nombre n'est écrit en texte", not textes_numeriques)

    # Un site qui n'a pas répondu doit laisser des cases VIDES. S'il en sortait
    # des zéros, « je ne sais pas » et « il n'y en a pas » deviendraient le même
    # chiffre, et personne ne pourrait plus les séparer.
    sans_reponse = [l for l in f.iter_rows(min_row=2, values_only=True)
                    if l and l[3] in ("Pas de réponse", "Clos sans réponse")]
    dit("un site sans réponse laisse des cases vides, jamais des zéros",
        bool(sans_reponse) and all(v is None for l in sans_reponse for v in l[4:-1]))

    d = wb["Définitions"]
    dit("le dictionnaire dit ce qu'on compte et ce qu'on ne compte pas",
        [c.value for c in d[1]][:4] == ["Clé", "Rubrique", "Indicateur", "Unité"])

    print(f"\n  {meta['campagne']} — {len(meta['onglets'])} onglets, "
          f"{sum(meta['lignes'])} lignes")
    shutil.rmtree(tmp, ignore_errors=True)
    return 1 if rates else 0

sys.exit(main())
